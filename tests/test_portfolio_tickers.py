import pandas as pd

from app.services.portfolio_tickers import load_portfolio_tickers


def test_missing_file(tmp_path, monkeypatch):
    monkeypatch.setenv("NEWS_ARCHIVE_ROOT", str(tmp_path))
    assert load_portfolio_tickers() == {}


def test_parses_xlsx_with_preamble(tmp_path, monkeypatch):
    monkeypatch.setenv("NEWS_ARCHIVE_ROOT", str(tmp_path))
    # Mimic the real xlsx: 3 preamble rows, then header, then data.
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Portfolio preamble"
    ws["A2"] = "Data as of: ..."
    for col_idx, col_name in enumerate(
        ["Account", "Holding", "Type", "Value (DKK)", "Acct Weight", "Total Weight", "Sector/Theme"],
        start=1,
    ):
        ws.cell(row=4, column=col_idx, value=col_name)
    ws.cell(row=5, column=1, value="ETORO")
    ws.cell(row=6, column=1, value="GOOG")
    ws.cell(row=6, column=7, value="US Tech")
    ws.cell(row=7, column=1, value="NVO")
    ws.cell(row=7, column=7, value="DK Pharma")
    wb.save(tmp_path / "Portfolio_Total_Weights.xlsx")
    result = load_portfolio_tickers()
    assert "GOOG" in result
    assert result["GOOG"] == "US Tech"
    assert result["NVO"] == "DK Pharma"
    assert "ETORO" not in result


def test_malformed_returns_empty(tmp_path, monkeypatch):
    monkeypatch.setenv("NEWS_ARCHIVE_ROOT", str(tmp_path))
    (tmp_path / "Portfolio_Total_Weights.xlsx").write_text("not a real xlsx")
    assert load_portfolio_tickers() == {}


def test_reads_real_file():
    """Smoke test against the user's actual xlsx — skip if not present."""
    import os

    from pathlib import Path

    root = Path(os.getenv("NEWS_ARCHIVE_ROOT", "/Users/simonbaumgart/Documents/Claude/Projects/Investment Ideas and Portfolio"))
    if not (root / "Portfolio_Total_Weights.xlsx").exists():
        return  # skip silently
    result = load_portfolio_tickers()
    assert len(result) > 5
    assert "GOOG" in result
